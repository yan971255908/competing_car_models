import io
import re
from typing import Any, Optional

from fastapi import HTTPException
from openpyxl import load_workbook

from app.models.sql import SourceDocumentType, TechnologyCategory, TechnologyMaturityLevel

SUPPORTED_FIELDS = [
    {"key": "brand_name", "label": "品牌", "group": "车型"},
    {"key": "model_name", "label": "车型", "group": "车型"},
    {"key": "variant_name", "label": "版本", "group": "车型"},
    {"key": "energy_type", "label": "能源类型", "group": "车型"},
    {"key": "market_segment", "label": "市场级别", "group": "车型"},
    {"key": "launch_year", "label": "上市年份", "group": "车型"},
    {"key": "base_price", "label": "指导价", "group": "车型"},
    {"key": "technology_name", "label": "技术点名称", "group": "技术点"},
    {"key": "technology_category", "label": "技术类别", "group": "技术点"},
    {"key": "technology_description", "label": "技术描述", "group": "技术点"},
    {"key": "maturity_level", "label": "成熟度", "group": "技术点"},
    {"key": "evidence_text", "label": "证据文本", "group": "证据"},
    {"key": "source_type", "label": "来源类型", "group": "证据"},
    {"key": "page_or_time", "label": "页码或时间", "group": "证据"},
    {"key": "confidence", "label": "置信度", "group": "证据"},
]
SUPPORTED_FIELD_KEYS = {item["key"] for item in SUPPORTED_FIELDS}
FIELD_ALIASES = {
    "brand_name": ("品牌", "厂商", "车企", "品牌名称", "brand", "brand_name"),
    "model_name": ("车型", "车系", "车型名称", "车款名称", "竞品车型", "车辆名称", "model", "model_name"),
    "variant_name": ("版本", "配置", "配置版本", "车型版本", "车款版本", "配置款", "variant", "version", "variant_name"),
    "energy_type": ("能源类型", "动力类型", "能源形式", "能源", "power_type", "energy_type"),
    "market_segment": ("级别", "车型级别", "市场级别", "细分市场", "segment", "market_segment"),
    "launch_year": ("上市年份", "年份", "年款", "launch_year"),
    "base_price": ("指导价", "官方指导价", "售价", "价格", "price", "base_price"),
    "technology_name": ("技术点", "技术点名称", "技术名称", "技术亮点", "核心技术", "technology", "technology_name"),
    "technology_category": ("技术类别", "技术分类", "类别", "category", "technology_category"),
    "technology_description": ("技术描述", "描述", "说明", "技术说明", "description", "technology_description"),
    "maturity_level": ("成熟度", "技术成熟度", "maturity_level"),
    "evidence_text": ("证据", "证据文本", "证据描述", "来源描述", "原文", "evidence", "evidence_text"),
    "source_type": ("来源类型", "数据来源", "来源", "source", "source_type"),
    "page_or_time": ("页码", "时间", "页码或时间", "位置", "时间点", "page", "page_or_time"),
    "confidence": ("置信度", "可信度", "confidence"),
}
CATEGORY_ALIASES = {
    "动力": TechnologyCategory.POWER, "power": TechnologyCategory.POWER,
    "底盘": TechnologyCategory.CHASSIS, "chassis": TechnologyCategory.CHASSIS,
    "智驾": TechnologyCategory.ADAS, "自动驾驶": TechnologyCategory.ADAS, "adas": TechnologyCategory.ADAS,
    "座舱": TechnologyCategory.COCKPIT, "cockpit": TechnologyCategory.COCKPIT,
    "电池": TechnologyCategory.BATTERY, "battery": TechnologyCategory.BATTERY,
    "电子电气": TechnologyCategory.EE_ARCHITECTURE, "电子电气架构": TechnologyCategory.EE_ARCHITECTURE,
    "eearchitecture": TechnologyCategory.EE_ARCHITECTURE,
    "车身": TechnologyCategory.BODY, "body": TechnologyCategory.BODY,
    "其他": TechnologyCategory.OTHER, "other": TechnologyCategory.OTHER,
}
MATURITY_ALIASES = {
    "概念": TechnologyMaturityLevel.CONCEPT, "concept": TechnologyMaturityLevel.CONCEPT,
    "已发布": TechnologyMaturityLevel.ANNOUNCED, "发布": TechnologyMaturityLevel.ANNOUNCED,
    "announced": TechnologyMaturityLevel.ANNOUNCED,
    "量产": TechnologyMaturityLevel.MASS_PRODUCTION, "已量产": TechnologyMaturityLevel.MASS_PRODUCTION,
    "massproduction": TechnologyMaturityLevel.MASS_PRODUCTION,
}
SOURCE_TYPE_ALIASES = {
    "excel": SourceDocumentType.EXCEL, "excel配置表": SourceDocumentType.EXCEL, "配置表": SourceDocumentType.EXCEL,
    "发布会稿件": SourceDocumentType.PRESS_RELEASE, "pressrelease": SourceDocumentType.PRESS_RELEASE,
    "官网新闻": SourceDocumentType.OFFICIAL_SITE, "officialsite": SourceDocumentType.OFFICIAL_SITE,
    "网页文本": SourceDocumentType.WEBPAGE, "webpage": SourceDocumentType.WEBPAGE,
    "视频转写": SourceDocumentType.TRANSCRIPT, "transcript": SourceDocumentType.TRANSCRIPT,
    "手工录入": SourceDocumentType.MANUAL, "manual": SourceDocumentType.MANUAL,
}
VEHICLE_AUXILIARY_FIELDS = {"brand_name", "variant_name", "energy_type", "market_segment", "launch_year", "base_price"}
TECHNOLOGY_AUXILIARY_FIELDS = {"technology_category", "technology_description", "maturity_level"}


def normalize_name(value: Any) -> str:
    return re.sub(r"[\s_\-（）()]+", "", str(value or "").strip().lower())


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text_value = str(value).strip()
    return text_value or None


def parse_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group()) if match else None


def normalize_confidence(value: Any) -> float:
    number = parse_float(value)
    if number is None:
        return 0.8
    if number > 1:
        number /= 100
    return max(0.0, min(1.0, number))


def detect_fields(headers: list[str]) -> dict[str, str]:
    aliases = {field: {normalize_name(alias) for alias in values} for field, values in FIELD_ALIASES.items()}
    mapping: dict[str, str] = {}
    mapped_fields: set[str] = set()
    for header in headers:
        normalized = normalize_name(header)
        for field, field_aliases in aliases.items():
            if field not in mapped_fields and normalized in field_aliases:
                mapping[header] = field
                mapped_fields.add(field)
                break
    return mapping


def validate_final_mapping(raw_headers: list[str], final_mapping: dict[str, Optional[str]]) -> dict[str, str]:
    cleaned: dict[str, str] = {}
    used_fields: set[str] = set()
    for header, field in final_mapping.items():
        if header not in raw_headers or not field:
            continue
        if field not in SUPPORTED_FIELD_KEYS:
            raise HTTPException(status_code=400, detail=f"不支持的标准字段: {field}")
        if field in used_fields:
            raise HTTPException(status_code=400, detail=f"标准字段重复映射: {field}")
        cleaned[header] = field
        used_fields.add(field)
    return cleaned


def get_missing_required_fields(mapping: dict[str, str]) -> list[str]:
    fields = set(mapping.values())
    missing = []
    if fields & VEHICLE_AUXILIARY_FIELDS and "model_name" not in fields:
        missing.append("model_name")
    if fields & TECHNOLOGY_AUXILIARY_FIELDS and "technology_name" not in fields:
        missing.append("technology_name")
    return missing


def parse_mapped_row(row_number: int, raw_row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    values = {field: raw_row.get(header) for header, field in mapping.items()}
    category = CATEGORY_ALIASES.get(normalize_name(values.get("technology_category")), TechnologyCategory.OTHER)
    maturity = MATURITY_ALIASES.get(normalize_name(values.get("maturity_level")), TechnologyMaturityLevel.CONCEPT)
    source_type = SOURCE_TYPE_ALIASES.get(normalize_name(values.get("source_type")), SourceDocumentType.EXCEL)
    launch_year = parse_float(values.get("launch_year"))
    return {
        "row_number": row_number,
        "brand_name": clean_text(values.get("brand_name")), "model_name": clean_text(values.get("model_name")),
        "variant_name": clean_text(values.get("variant_name")), "energy_type": clean_text(values.get("energy_type")),
        "market_segment": clean_text(values.get("market_segment")),
        "launch_year": int(launch_year) if launch_year is not None else None,
        "base_price": parse_float(values.get("base_price")),
        "technology_name": clean_text(values.get("technology_name")), "technology_category": category.value,
        "technology_description": clean_text(values.get("technology_description")), "maturity_level": maturity.value,
        "evidence_text": clean_text(values.get("evidence_text")), "source_type": source_type.value,
        "page_or_time": clean_text(values.get("page_or_time")), "confidence": normalize_confidence(values.get("confidence")),
    }


def transform_rows(
    raw_headers: list[str],
    raw_rows: list[dict[str, Any]],
    final_mapping: dict[str, Optional[str]],
    enforce_required: bool = True,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    mapping = validate_final_mapping(raw_headers, final_mapping)
    missing = get_missing_required_fields(mapping)
    if enforce_required and "model_name" in missing:
        raise HTTPException(status_code=400, detail="缺少车型字段，无法导入车型数据")
    if enforce_required and "technology_name" in missing:
        raise HTTPException(status_code=400, detail="缺少技术点名称，无法导入技术点数据")
    rows = []
    for position, raw_row in enumerate(raw_rows, start=2):
        row_number = int(raw_row.get("__row_number__") or position)
        parsed = parse_mapped_row(row_number, raw_row, mapping)
        if parsed["model_name"] or parsed["technology_name"] or parsed["evidence_text"]:
            rows.append(parsed)
    return rows, mapping


def summarize_rows(rows: list[dict[str, Any]]) -> dict[str, int]:
    vehicles = {((row["brand_name"] or "未填写品牌").casefold(), row["model_name"].casefold()) for row in rows if row["model_name"]}
    technologies = {(row["technology_name"].casefold(), row["technology_category"]) for row in rows if row["technology_name"]}
    evidence = {row["evidence_text"] for row in rows if row["evidence_text"]}
    return {"row_count": len(rows), "vehicle_count": len(vehicles), "technology_count": len(technologies), "evidence_count": len(evidence)}


def build_preview(content: bytes, file_name: str) -> dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel 文件无法解析: {exc}") from exc
    worksheet = workbook.worksheets[0]
    iterator = worksheet.iter_rows(values_only=True)
    raw_headers = [clean_text(value) or "" for value in list(next(iterator, ()))]
    if not any(raw_headers):
        raise HTTPException(status_code=400, detail="Excel 第一行必须包含表头")
    nonempty_headers = [header for header in raw_headers if header]
    if len(nonempty_headers) != len(set(nonempty_headers)):
        raise HTTPException(status_code=400, detail="Excel 表头存在重复列名，请先调整后重试")
    auto_mapping = detect_fields(raw_headers)
    raw_rows = []
    for row_number, values in enumerate(iterator, start=2):
        if not any(clean_text(value) for value in values):
            continue
        raw_row = {header: values[index] if index < len(values) else None for index, header in enumerate(raw_headers) if header}
        raw_row["__row_number__"] = row_number
        raw_rows.append(raw_row)
    transformed_rows, _ = transform_rows(raw_headers, raw_rows, auto_mapping, enforce_required=False)
    missing = get_missing_required_fields(auto_mapping)
    return {
        "file_name": file_name, "sheet_name": worksheet.title,
        "raw_headers": raw_headers, "auto_mapping": auto_mapping,
        "unmapped_headers": [header for header in raw_headers if header and header not in auto_mapping],
        "supported_fields": SUPPORTED_FIELDS, "missing_required_fields": missing,
        "preview_rows": raw_rows[:20], "raw_rows": raw_rows, "summary": summarize_rows(transformed_rows),
        "headers": raw_headers, "field_mapping": auto_mapping, "missing_fields": missing, "rows": transformed_rows,
    }
